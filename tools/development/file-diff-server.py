#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能文件对比工具 - 后端服务
支持高级功能：文件夹对比、版本历史、协作等
"""

import os
import json
import hashlib
import difflib
import mimetypes
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import sqlite3
import tempfile
import zipfile
import shutil

# 文档解析库
try:
    import PyPDF2
    import docx
    import openpyxl
    import markdown
    from bs4 import BeautifulSoup
except ImportError:
    print("警告: 某些文档解析库未安装，部分功能可能不可用")

class FileParser:
    """文件解析器类"""
    
    @staticmethod
    def get_file_type(file_path: str) -> str:
        """获取文件类型"""
        mime_type, _ = mimetypes.guess_type(file_path)
        extension = Path(file_path).suffix.lower()
        
        if extension in ['.pdf']:
            return 'pdf'
        elif extension in ['.doc', '.docx']:
            return 'word'
        elif extension in ['.xls', '.xlsx']:
            return 'excel'
        elif extension in ['.md', '.markdown']:
            return 'markdown'
        elif extension in ['.txt', '.log']:
            return 'text'
        elif extension in ['.py', '.js', '.java', '.cpp', '.c', '.h', '.css', '.html', '.xml', '.json']:
            return 'code'
        else:
            return 'unknown'
    
    @staticmethod
    def parse_file(file_path: str) -> Dict[str, Any]:
        """解析文件内容"""
        file_type = FileParser.get_file_type(file_path)
        file_size = os.path.getsize(file_path)
        file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
        
        try:
            if file_type == 'pdf':
                content = FileParser._parse_pdf(file_path)
            elif file_type == 'word':
                content = FileParser._parse_word(file_path)
            elif file_type == 'excel':
                content = FileParser._parse_excel(file_path)
            elif file_type == 'markdown':
                content = FileParser._parse_markdown(file_path)
            else:
                content = FileParser._parse_text(file_path)
                
            return {
                'success': True,
                'content': content,
                'file_type': file_type,
                'file_size': file_size,
                'file_modified': file_modified.isoformat(),
                'line_count': len(content.split('\n')) if content else 0,
                'char_count': len(content) if content else 0
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'file_type': file_type,
                'file_size': file_size,
                'file_modified': file_modified.isoformat()
            }
    
    @staticmethod
    def _parse_pdf(file_path: str) -> str:
        """解析PDF文件"""
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except:
            return FileParser._parse_text(file_path)
    
    @staticmethod
    def _parse_word(file_path: str) -> str:
        """解析Word文档"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except:
            return FileParser._parse_text(file_path)
    
    @staticmethod
    def _parse_excel(file_path: str) -> str:
        """解析Excel文件"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"=== {sheet_name} ===\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
                text += "\n"
            return text
        except:
            return FileParser._parse_text(file_path)
    
    @staticmethod
    def _parse_markdown(file_path: str) -> str:
        """解析Markdown文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                # 保持原始markdown格式
                return content
        except:
            return FileParser._parse_text(file_path)
    
    @staticmethod
    def _parse_text(file_path: str) -> str:
        """解析文本文件"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    return file.read()
            except UnicodeDecodeError:
                continue
        raise Exception("无法解析文件编码")

class DiffEngine:
    """高级差异对比引擎"""
    
    @staticmethod
    def compare_texts(text1: str, text2: str, mode: str = 'line', **options) -> Dict[str, Any]:
        """文本对比"""
        # 预处理选项
        ignore_whitespace = options.get('ignore_whitespace', False)
        ignore_case = options.get('ignore_case', False)
        context_lines = options.get('context_lines', 3)
        
        # 预处理文本
        processed_text1 = DiffEngine._preprocess_text(text1, ignore_whitespace, ignore_case)
        processed_text2 = DiffEngine._preprocess_text(text2, ignore_whitespace, ignore_case)
        
        # 执行对比
        if mode == 'line':
            diff = list(difflib.unified_diff(
                processed_text1.splitlines(keepends=True),
                processed_text2.splitlines(keepends=True),
                fromfile='文件1',
                tofile='文件2',
                n=context_lines
            ))
        elif mode == 'word':
            diff = DiffEngine._word_diff(processed_text1, processed_text2)
        elif mode == 'char':
            diff = DiffEngine._char_diff(processed_text1, processed_text2)
        else:
            diff = DiffEngine._semantic_diff(processed_text1, processed_text2)
        
        # 计算统计信息
        stats = DiffEngine._calculate_stats(text1, text2, diff, mode)
        
        return {
            'diff': diff,
            'stats': stats,
            'mode': mode,
            'options': options
        }
    
    @staticmethod
    def _preprocess_text(text: str, ignore_whitespace: bool, ignore_case: bool) -> str:
        """预处理文本"""
        if ignore_whitespace:
            # 标准化空白字符
            import re
            text = re.sub(r'\s+', ' ', text).strip()
        
        if ignore_case:
            text = text.lower()
        
        return text
    
    @staticmethod
    def _word_diff(text1: str, text2: str) -> List[str]:
        """单词级别对比"""
        words1 = text1.split()
        words2 = text2.split()
        
        return list(difflib.unified_diff(words1, words2, lineterm=''))
    
    @staticmethod
    def _char_diff(text1: str, text2: str) -> List[str]:
        """字符级别对比"""
        return list(difflib.unified_diff(list(text1), list(text2), lineterm=''))
    
    @staticmethod
    def _semantic_diff(text1: str, text2: str) -> List[str]:
        """语义级别对比（句子级别）"""
        import re
        
        # 按句子分割
        sentences1 = re.split(r'[.!?。！？]\s*', text1)
        sentences2 = re.split(r'[.!?。！？]\s*', text2)
        
        return list(difflib.unified_diff(sentences1, sentences2, lineterm=''))
    
    @staticmethod
    def _calculate_stats(text1: str, text2: str, diff: List[str], mode: str) -> Dict[str, Any]:
        """计算统计信息"""
        lines1 = text1.splitlines()
        lines2 = text2.splitlines()
        
        added = sum(1 for line in diff if line.startswith('+') and not line.startswith('+++'))
        removed = sum(1 for line in diff if line.startswith('-') and not line.startswith('---'))
        
        total_lines1 = len(lines1)
        total_lines2 = len(lines2)
        
        # 计算相似度（基于最长公共子序列）
        similarity = DiffEngine._calculate_similarity(text1, text2)
        
        return {
            'added': added,
            'removed': removed,
            'modified': min(added, removed),
            'total_changes': added + removed,
            'total_lines1': total_lines1,
            'total_lines2': total_lines2,
            'similarity': round(similarity * 100, 2),
            'mode': mode
        }
    
    @staticmethod
    def _calculate_similarity(text1: str, text2: str) -> float:
        """计算文本相似度"""
        sequence_matcher = difflib.SequenceMatcher(None, text1, text2)
        return sequence_matcher.ratio()

class FolderComparator:
    """文件夹对比器"""
    
    @staticmethod
    def compare_folders(folder1: str, folder2: str, **options) -> Dict[str, Any]:
        """对比两个文件夹"""
        include_patterns = options.get('include_patterns', ['*'])
        exclude_patterns = options.get('exclude_patterns', [])
        recursive = options.get('recursive', True)
        
        # 获取文件列表
        files1 = FolderComparator._get_file_list(folder1, include_patterns, exclude_patterns, recursive)
        files2 = FolderComparator._get_file_list(folder2, include_patterns, exclude_patterns, recursive)
        
        # 分类文件
        only_in_1 = []
        only_in_2 = []
        common_files = []
        different_files = []
        
        all_files = set(files1.keys()) | set(files2.keys())
        
        for relative_path in all_files:
            if relative_path in files1 and relative_path in files2:
                # 对比文件内容
                file1_info = files1[relative_path]
                file2_info = files2[relative_path]
                
                if FolderComparator._files_are_different(file1_info['path'], file2_info['path']):
                    different_files.append({
                        'path': relative_path,
                        'file1': file1_info,
                        'file2': file2_info
                    })
                else:
                    common_files.append({
                        'path': relative_path,
                        'file1': file1_info,
                        'file2': file2_info
                    })
            elif relative_path in files1:
                only_in_1.append({
                    'path': relative_path,
                    'file': files1[relative_path]
                })
            else:
                only_in_2.append({
                    'path': relative_path,
                    'file': files2[relative_path]
                })
        
        return {
            'folder1': folder1,
            'folder2': folder2,
            'only_in_1': only_in_1,
            'only_in_2': only_in_2,
            'common_files': common_files,
            'different_files': different_files,
            'summary': {
                'total_files_1': len(files1),
                'total_files_2': len(files2),
                'only_in_1_count': len(only_in_1),
                'only_in_2_count': len(only_in_2),
                'common_count': len(common_files),
                'different_count': len(different_files)
            }
        }
    
    @staticmethod
    def _get_file_list(folder: str, include_patterns: List[str], exclude_patterns: List[str], recursive: bool) -> Dict[str, Dict]:
        """获取文件夹中的文件列表"""
        import fnmatch
        
        files = {}
        folder_path = Path(folder)
        
        if recursive:
            pattern = '**/*'
        else:
            pattern = '*'
        
        for file_path in folder_path.glob(pattern):
            if file_path.is_file():
                relative_path = str(file_path.relative_to(folder_path))
                
                # 检查包含模式
                include_match = any(fnmatch.fnmatch(relative_path, pattern) for pattern in include_patterns)
                if not include_match:
                    continue
                
                # 检查排除模式
                exclude_match = any(fnmatch.fnmatch(relative_path, pattern) for pattern in exclude_patterns)
                if exclude_match:
                    continue
                
                files[relative_path] = {
                    'path': str(file_path),
                    'size': file_path.stat().st_size,
                    'modified': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                    'hash': FolderComparator._calculate_file_hash(str(file_path))
                }
        
        return files
    
    @staticmethod
    def _calculate_file_hash(file_path: str) -> str:
        """计算文件哈希值"""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except:
            return ""
    
    @staticmethod
    def _files_are_different(file1: str, file2: str) -> bool:
        """检查两个文件是否不同"""
        hash1 = FolderComparator._calculate_file_hash(file1)
        hash2 = FolderComparator._calculate_file_hash(file2)
        return hash1 != hash2

class VersionHistory:
    """版本历史管理"""
    
    def __init__(self, db_path: str = "file_diff_history.db"):
        self.db_path = db_path
        self._init_database()
    
    def _init_database(self):
        """初始化数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS comparisons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                file1_name TEXT NOT NULL,
                file2_name TEXT NOT NULL,
                file1_hash TEXT,
                file2_hash TEXT,
                diff_summary TEXT,
                stats TEXT,
                notes TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS file_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                content TEXT,
                timestamp TEXT NOT NULL,
                version INTEGER NOT NULL
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def save_comparison(self, file1_path: str, file2_path: str, diff_result: Dict, notes: str = "") -> int:
        """保存对比结果"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        file1_hash = FolderComparator._calculate_file_hash(file1_path)
        file2_hash = FolderComparator._calculate_file_hash(file2_path)
        
        cursor.execute('''
            INSERT INTO comparisons (timestamp, file1_name, file2_name, file1_hash, file2_hash, diff_summary, stats, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            datetime.now().isoformat(),
            os.path.basename(file1_path),
            os.path.basename(file2_path),
            file1_hash,
            file2_hash,
            json.dumps(diff_result.get('diff', [])[:10]),  # 只保存前10行差异
            json.dumps(diff_result.get('stats', {})),
            notes
        ))
        
        comparison_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return comparison_id
    
    def get_comparison_history(self, limit: int = 50) -> List[Dict]:
        """获取对比历史"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, timestamp, file1_name, file2_name, stats, notes
            FROM comparisons
            ORDER BY timestamp DESC
            LIMIT ?
        ''', (limit,))
        
        results = []
        for row in cursor.fetchall():
            results.append({
                'id': row[0],
                'timestamp': row[1],
                'file1_name': row[2],
                'file2_name': row[3],
                'stats': json.loads(row[4]) if row[4] else {},
                'notes': row[5]
            })
        
        conn.close()
        return results
    
    def save_file_version(self, file_path: str, content: str) -> int:
        """保存文件版本"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        file_hash = hashlib.md5(content.encode()).hexdigest()
        
        # 获取下一个版本号
        cursor.execute('SELECT MAX(version) FROM file_versions WHERE file_path = ?', (file_path,))
        result = cursor.fetchone()
        next_version = (result[0] or 0) + 1
        
        cursor.execute('''
            INSERT INTO file_versions (file_path, file_hash, content, timestamp, version)
            VALUES (?, ?, ?, ?, ?)
        ''', (file_path, file_hash, content, datetime.now().isoformat(), next_version))
        
        version_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return version_id
    
    def get_file_versions(self, file_path: str) -> List[Dict]:
        """获取文件版本历史"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, version, timestamp, file_hash
            FROM file_versions
            WHERE file_path = ?
            ORDER BY version DESC
        ''', (file_path,))
        
        results = []
        for row in cursor.fetchall():
            results.append({
                'id': row[0],
                'version': row[1],
                'timestamp': row[2],
                'file_hash': row[3]
            })
        
        conn.close()
        return results

class FileDiffAPI:
    """文件对比API类"""
    
    def __init__(self):
        self.history = VersionHistory()
    
    def compare_files(self, file1_path: str, file2_path: str, **options) -> Dict[str, Any]:
        """对比两个文件"""
        try:
            # 解析文件
            file1_result = FileParser.parse_file(file1_path)
            file2_result = FileParser.parse_file(file2_path)
            
            if not file1_result['success']:
                return {'success': False, 'error': f"文件1解析失败: {file1_result['error']}"}
            
            if not file2_result['success']:
                return {'success': False, 'error': f"文件2解析失败: {file2_result['error']}"}
            
            # 执行对比
            diff_result = DiffEngine.compare_texts(
                file1_result['content'],
                file2_result['content'],
                **options
            )
            
            # 保存到历史记录
            comparison_id = self.history.save_comparison(
                file1_path, file2_path, diff_result, options.get('notes', '')
            )
            
            return {
                'success': True,
                'comparison_id': comparison_id,
                'file1': file1_result,
                'file2': file2_result,
                'diff': diff_result,
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def compare_folders(self, folder1_path: str, folder2_path: str, **options) -> Dict[str, Any]:
        """对比两个文件夹"""
        try:
            result = FolderComparator.compare_folders(folder1_path, folder2_path, **options)
            return {'success': True, **result}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def get_history(self, limit: int = 50) -> Dict[str, Any]:
        """获取对比历史"""
        try:
            history = self.history.get_comparison_history(limit)
            return {'success': True, 'history': history}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def export_comparison(self, comparison_data: Dict, format: str = 'html') -> str:
        """导出对比结果"""
        if format == 'html':
            return self._export_to_html(comparison_data)
        elif format == 'json':
            return json.dumps(comparison_data, indent=2, ensure_ascii=False)
        elif format == 'text':
            return self._export_to_text(comparison_data)
        else:
            raise ValueError(f"不支持的导出格式: {format}")
    
    def _export_to_html(self, data: Dict) -> str:
        """导出为HTML格式"""
        html_template = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>文件对比结果</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .header { border-bottom: 2px solid #ccc; padding-bottom: 10px; margin-bottom: 20px; }
        .stats { display: flex; gap: 20px; margin-bottom: 20px; }
        .stat { text-align: center; padding: 10px; border: 1px solid #ddd; border-radius: 5px; }
        .diff-line { font-family: monospace; white-space: pre; }
        .added { background-color: #d4edda; }
        .removed { background-color: #f8d7da; }
    </style>
</head>
<body>
    <div class="header">
        <h1>文件对比结果</h1>
        <p>生成时间: {timestamp}</p>
        <p>文件1: {file1_name}</p>
        <p>文件2: {file2_name}</p>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div><strong>{total_changes}</strong></div>
            <div>总变更</div>
        </div>
        <div class="stat">
            <div><strong>{added}</strong></div>
            <div>新增</div>
        </div>
        <div class="stat">
            <div><strong>{removed}</strong></div>
            <div>删除</div>
        </div>
        <div class="stat">
            <div><strong>{similarity}%</strong></div>
            <div>相似度</div>
        </div>
    </div>
    
    <div class="diff-content">
        <h3>详细对比:</h3>
        <div style="font-family: monospace; white-space: pre-wrap;">{diff_content}</div>
    </div>
</body>
</html>
        """
        
        stats = data.get('diff', {}).get('stats', {})
        diff_lines = data.get('diff', {}).get('diff', [])
        
        return html_template.format(
            timestamp=data.get('timestamp', ''),
            file1_name=data.get('file1', {}).get('file_type', 'Unknown'),
            file2_name=data.get('file2', {}).get('file_type', 'Unknown'),
            total_changes=stats.get('total_changes', 0),
            added=stats.get('added', 0),
            removed=stats.get('removed', 0),
            similarity=stats.get('similarity', 0),
            diff_content='\n'.join(diff_lines)
        )
    
    def _export_to_text(self, data: Dict) -> str:
        """导出为文本格式"""
        lines = []
        lines.append("文件对比结果")
        lines.append("=" * 50)
        lines.append(f"生成时间: {data.get('timestamp', '')}")
        lines.append(f"文件1: {data.get('file1', {}).get('file_type', 'Unknown')}")
        lines.append(f"文件2: {data.get('file2', {}).get('file_type', 'Unknown')}")
        lines.append("")
        
        stats = data.get('diff', {}).get('stats', {})
        lines.append("统计信息:")
        lines.append(f"总变更: {stats.get('total_changes', 0)}")
        lines.append(f"新增: {stats.get('added', 0)}")
        lines.append(f"删除: {stats.get('removed', 0)}")
        lines.append(f"相似度: {stats.get('similarity', 0)}%")
        lines.append("")
        
        lines.append("详细对比:")
        lines.append("-" * 30)
        diff_lines = data.get('diff', {}).get('diff', [])
        lines.extend(diff_lines)
        
        return '\n'.join(lines)

# 命令行接口
def main():
    """命令行主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='智能文件对比工具')
    parser.add_argument('file1', help='第一个文件路径')
    parser.add_argument('file2', help='第二个文件路径')
    parser.add_argument('--mode', choices=['line', 'word', 'char', 'semantic'], 
                       default='line', help='对比模式')
    parser.add_argument('--ignore-whitespace', action='store_true', help='忽略空白字符')
    parser.add_argument('--ignore-case', action='store_true', help='忽略大小写')
    parser.add_argument('--output', help='输出文件路径')
    parser.add_argument('--format', choices=['html', 'json', 'text'], 
                       default='text', help='输出格式')
    
    args = parser.parse_args()
    
    # 创建API实例
    api = FileDiffAPI()
    
    # 执行对比
    result = api.compare_files(
        args.file1, 
        args.file2,
        mode=args.mode,
        ignore_whitespace=args.ignore_whitespace,
        ignore_case=args.ignore_case
    )
    
    if not result['success']:
        print(f"错误: {result['error']}")
        return 1
    
    # 导出结果
    output_content = api.export_comparison(result, args.format)
    
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output_content)
        print(f"结果已保存到: {args.output}")
    else:
        print(output_content)
    
    return 0

if __name__ == '__main__':
    exit(main())